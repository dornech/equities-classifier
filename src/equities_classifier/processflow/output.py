"""Excel output handling for the security process flow."""


# ruff and mypy per file settings
#
# boolean-type arguments
# ruff: noqa: FBT001, FBT002
# others
# ruff: noqa: RUF105


# fmt: off


from typing import Any
from collections.abc import Sequence

from enum import Enum
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from equities_classifier.enums import SecurityIdentifierType
from equities_classifier.models import Security, ClassificationSystem
from equities_classifier.classification.systems import GECS, GICS


class ClassificationOutput(Enum):
    """Classification systems to include in the output."""

    NONE = "none"
    GECS = "gecs"
    GICS = "gics"
    ALL = "all"


def get_classification_output(
    gecs: bool = False,
    gics: bool = False,
) -> ClassificationOutput:
    """Return selected classification output."""

    if gecs and gics:
        return ClassificationOutput.ALL
    if gecs:
        return ClassificationOutput.GECS
    if gics:
        return ClassificationOutput.GICS

    return ClassificationOutput.NONE


def write_excel(
    securities: list[Security],
    filename: str | Path,
    *,
    classifications: ClassificationOutput = ClassificationOutput.NONE,
    provider_details: bool = False,
) -> None:
    """Write securities to an Excel file."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Securities"

    columns = _create_columns(
        securities,
        classifications,
        provider_details,
    )
    for column_index, column_name in enumerate(columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, security in enumerate(securities, start=2):
        values = _create_row(
            security,
            columns,
            classifications,
            provider_details,
        )
        for column_index, value in enumerate(values, start=1):
            if value and isinstance(value, list):
                value = ", ".join(value) if len(value) > 0 else ""
            worksheet.cell(row=row_index, column=column_index, value=value)

    _format_worksheet(worksheet)

    workbook.save(filename)


def _create_columns(
    securities: list[Security],
    classifications: ClassificationOutput,
    provider_details: bool,
) -> list[str]:
    """Create output columns."""

    columns = ["Name", "Ticker"]

    identifier_types = _get_identifier_types(securities)
    columns.extend(identifier_type.value.upper() for identifier_type in identifier_types)

    if classifications in {ClassificationOutput.GECS, ClassificationOutput.ALL}:
        columns.extend(_classification_columns(GECS))
    if classifications in {ClassificationOutput.GICS, ClassificationOutput.ALL}:
        columns.extend(_classification_columns(GICS))

    if provider_details:
        columns.extend(_provider_columns(securities))

    return columns


def _get_identifier_types(
    securities: list[Security],
) -> list[SecurityIdentifierType]:
    """Return identifier types occurring in the securities."""

    return sorted(
        {
            identifier.type
            for security in securities
            for identifier in security.identifiers
        },
        key=lambda identifier_type: identifier_type.value,
    )


def _classification_columns(
    system: ClassificationSystem,
) -> list[str]:
    """Return classification columns for a classification system."""

    columns: list[str] = []

    for level_name in system.hierarchy:
        column_name = f"{system.short_name} {level_name}"
        if system.supports_codes:
            columns.append(f"{column_name} Code")
        columns.append(column_name)

    return columns


def _provider_columns(
    securities: list[Security],
) -> list[str]:
    """Return provider attribute columns."""

    columns: list[str] = []

    for security in securities:
        for datasource, attributes in security.provider_attributes.items():
            for attribute in attributes:
                column = f"{datasource}.{attribute}"
                if column not in columns:
                    columns.append(column)

    return sorted(columns)


def _create_row(
    security: Security,
    columns: Sequence[str],
    classifications: ClassificationOutput,
    provider_details: bool,
) -> list[Any]:
    """Create one output row using the output columns as keys."""

    values: dict[str, Any] = {"Name": security.name, "Ticker": security.ticker}

    for identifier in security.identifiers:
        values[identifier.type.value.upper()] = identifier.value

    if classifications in {ClassificationOutput.GECS, ClassificationOutput.ALL}:
        _add_classification_values(values, security, GECS)
    if classifications in {ClassificationOutput.GICS, ClassificationOutput.ALL}:
        _add_classification_values(values, security, GICS)

    if provider_details:
        _add_provider_values(values, security)

    return [
        values.get(column)
        for column in columns
    ]


def _add_classification_values(
    values: dict[str, Any],
    security: Security,
    system: ClassificationSystem,
) -> None:
    """Add classification values to an output row."""

    for classification in security.classifications:
        if classification.system.id != system.id:
            continue

        for level_name, node in zip(
            system.hierarchy,
            classification.nodes.values(),
            strict=True,
        ):
            column_name = f"{system.short_name} {level_name}"
            if system.supports_codes:
                values[f"{column_name} Code"] = node.code
            values[column_name] = node.value


def _add_provider_values(
    values: dict[str, Any],
    security: Security,
) -> None:
    """Add provider attributes to an output row."""

    for datasource, attributes in security.provider_attributes.items():
        for attribute, value in attributes.items():
            if isinstance(value, list):
                value = ", ".join(map(str, value)) if len(value) > 0 else ""
            values[f"{datasource}.{attribute}"] = value


def _format_worksheet(
    worksheet: Any,
) -> None:
    """Format Excel worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        column_letter = get_column_letter(column_cells[0].column)

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)
