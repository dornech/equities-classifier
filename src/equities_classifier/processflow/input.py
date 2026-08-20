"""Input handling for the security classification process."""


# ruff and mypy per file settings
#
# others
# ruff: noqa: PLR2004, RUF105
#
# fmt: off


from pathlib import Path

from openpyxl import load_workbook

from equities_classifier.models import (
    SecurityIdentifier,
    SecurityIdentifierType,
)


def read_identifiers(
    filename: str | Path,
) -> list[SecurityIdentifier]:
    """Read security identifiers from an Excel file.

    The input file must contain the columns ``Type`` and ``Value``.
    ``Type`` may be omitted or empty. In that case the identifier type
    is detected automatically for ISINs and tickers.
    """

    path = Path(filename)
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)

        headers = next(rows, None)
        if headers is None:
            return []

        header_map = _get_header_map(headers)
        if "value" not in header_map:
            message = "Input file must contain a 'Value' column."
            raise ValueError(message)

        return [
            _create_identifier(row, header_map, row_number)
            for row_number, row in enumerate(rows, start=2)
        ]
    finally:
        workbook.close()


def _get_header_map(headers: tuple[object, ...]) -> dict[str, int]:
    """Create a normalized header-to-column mapping."""

    return {
        str(header).strip().casefold(): index
        for index, header in enumerate(headers)
        if header is not None
    }


def _create_identifier(
    row: tuple[object, ...],
    header_map: dict[str, int],
    row_number: int,
) -> SecurityIdentifier:
    """Create a SecurityIdentifier from an input row."""

    value = _get_cell_value(row, header_map["value"])
    if not value:
        message = f"Missing identifier value in row {row_number}."
        raise ValueError(message)

    type_index = header_map.get("type")
    identifier_type = None
    if type_index is not None:
        type_value = _get_cell_value(row, type_index)
        if type_value:
            identifier_type = _parse_identifier_type(type_value, row_number)

    if identifier_type is None:
        identifier_type = _detect_identifier_type(value)

    if identifier_type is None:
        message = f"Unable to determine identifier type for {value!r} in row {row_number}."
        raise ValueError(message)

    return SecurityIdentifier(type=identifier_type, value=value)


def _get_cell_value(
    row: tuple[object, ...],
    index: int,
) -> str:
    """Return a normalized string cell value."""

    if index >= len(row) or row[index] is None:
        return ""

    return str(row[index]).strip().upper()


def _parse_identifier_type(
    value: str,
    row_number: int,
) -> SecurityIdentifierType:
    """Parse an explicitly supplied identifier type."""

    normalized = value.strip().lower()

    try:
        return SecurityIdentifierType(normalized)
    except ValueError as exc:
        message = f"Unknown security identifier type  {value!r} in row {row_number}."
        raise ValueError(message) from exc


def _detect_identifier_type(
    value: str,
) -> SecurityIdentifierType | None:
    """Detect an identifier type from its value."""

    if _is_valid_isin(value):
        return SecurityIdentifierType.ISIN
    if value.isalpha():
        return SecurityIdentifierType.TICKER

    return None


def _is_valid_isin(value: str) -> bool:
    """Return whether a value is a valid ISIN."""

    if len(value) != 12:
        return False
    if not value[:2].isalpha():
        return False
    if not value[2:].isalnum():
        return False

    converted = "".join(
        str(ord(character.upper()) - ord("A") + 10)
        if character.isalpha()
        else character
        for character in value
    )

    # apply Luhn algorithm - https://en.wikipedia.org/wiki/Luhn_algorithm
    digits = [int(digit) for digit in converted]
    total = 0
    for index, digit in enumerate(reversed(digits)):
        if index % 2 == 1:
            digit *= 2
            if digit > 9:
                digit -= 9
        total += digit

    return total % 10 == 0
