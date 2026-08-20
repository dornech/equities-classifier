"""Tests for processflow output handling."""


# ruff and mypy per file settings
#
# others
# ruff: noqa: RUF105

# fmt: off


from pathlib import Path

import openpyxl

from equities_classifier.enums import (
    ClassificationLevel,
    SecurityIdentifierType,
)
from equities_classifier.models import (
    ClassificationNode,
    SecurityIdentifier,
    Security,
    SecurityClassification,
)
from equities_classifier.classification.systems import (
    GECS,
    GICS,
)
from equities_classifier.processflow.output import (
    ClassificationOutput,
    write_excel,
)


def _load_worksheet(
    filename: Path,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Load the output worksheet."""

    workbook = openpyxl.load_workbook(filename)
    return workbook["Securities"]


def _headers(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> list[str]:
    """Return worksheet headers."""

    return [
        cell.value
        for cell in worksheet[1]
    ]


def _security(
    *,
    name: str = "Apple Inc.",
    ticker: str = "AAPL",
) -> Security:
    """Create a test security."""

    return Security(
        name=name,
        ticker=ticker,
    )


def test_write_excel_basic(
    tmp_path: Path,
) -> None:
    """Test basic Excel output."""

    filename = tmp_path / "output.xlsx"

    security = _security()
    security.identifiers.append(SecurityIdentifier(type=SecurityIdentifierType.ISIN, value="US0378331005"))

    write_excel([security], filename,)

    worksheet = _load_worksheet(filename)
    headers = _headers(worksheet)

    assert headers == ["Name", "Ticker", "ISIN",]

    assert worksheet.cell(2, 1).value == "Apple Inc."
    assert worksheet.cell(2, 2).value == "AAPL"
    assert worksheet.cell(2, 3).value == "US0378331005"

    worksheet.parent.close()


def test_write_excel_identifier_union(
    tmp_path: Path,
) -> None:
    """Test identifier columns are the union of all identifiers."""

    filename = tmp_path / "output.xlsx"

    apple = Security(name="Apple Inc.", ticker="AAPL")
    apple.identifiers.append(SecurityIdentifier(type=SecurityIdentifierType.ISIN, value="US0378331005"))

    microsoft = Security(name="Microsoft Corporation", ticker="MSFT")
    microsoft.identifiers.append(SecurityIdentifier(type=SecurityIdentifierType.ISIN, value="US5949181045"))
    microsoft.identifiers.append(SecurityIdentifier(type=SecurityIdentifierType.WKN, value="870747"))

    write_excel([apple, microsoft], filename)

    worksheet = _load_worksheet(filename)

    headers = [cell.value for cell in worksheet[1]]

    assert headers[:4] == ["Name", "Ticker", "ISIN", "WKN"]

    # Apple has no WKN.
    assert worksheet["D2"].value is None

    # Microsoft has a WKN.
    assert worksheet["D3"].value == "870747"

    worksheet.parent.close()


def test_write_excel_provider_attribute_union(
    tmp_path: Path,
) -> None:
    """Test provider columns are the union of all attributes."""

    filename = tmp_path / "output.xlsx"

    apple = Security(name="Apple Inc.", ticker="AAPL",)
    apple.provider_attributes["morningstar"] = {"sector": "Technology", "industry": "Consumer Electronics"}

    microsoft = Security(name="Microsoft Corporation", ticker="MSFT")
    microsoft.provider_attributes["morningstar"] = {"sector": "Technology"}
    microsoft.provider_attributes["openfigi"] = {"figi": "BBG000BPH459"}

    write_excel([apple, microsoft], filename, provider_details=True)

    worksheet = _load_worksheet(filename)

    headers = [cell.value for cell in worksheet[1]]

    assert "morningstar.sector" in headers
    assert "morningstar.industry" in headers
    assert "openfigi.figi" in headers

    sector_column = headers.index("morningstar.sector") + 1
    industry_column = headers.index("morningstar.industry") + 1
    figi_column = headers.index("openfigi.figi") + 1

    assert worksheet.cell(2, sector_column).value == "Technology"
    assert worksheet.cell(2, industry_column).value == "Consumer Electronics"
    assert worksheet.cell(2, figi_column).value is None

    assert worksheet.cell(3, sector_column).value == "Technology"
    assert worksheet.cell(3, industry_column).value is None
    assert worksheet.cell(3, figi_column).value == "BBG000BPH459"

    worksheet.parent.close()


def test_write_excel_gecs(
    tmp_path: Path,
) -> None:
    """Test GECS classification output."""

    filename = tmp_path / "output.xlsx"

    security = _security()
    security.classifications.append(
        SecurityClassification(
            system=GECS,
            nodes={
                ClassificationLevel.LEVEL1: ClassificationNode(value="Sensitive",),
                ClassificationLevel.LEVEL2: ClassificationNode(value="Technology",),
                ClassificationLevel.LEVEL3: ClassificationNode(value="Software",),
            },
        )
    )

    write_excel([security], filename, classifications=ClassificationOutput.GECS,)

    worksheet = _load_worksheet(filename)

    assert _headers(worksheet) == ["Name", "Ticker", "GECS Super Sector", "GECS Sector", "GECS Industry",]

    assert worksheet.cell(2, 3).value == "Sensitive"
    assert worksheet.cell(2, 4).value == "Technology"
    assert worksheet.cell(2, 5).value == "Software"

    worksheet.parent.close()


def test_write_excel_gics(
    tmp_path: Path,
) -> None:
    """Test GICS classification output including codes."""

    filename = tmp_path / "output.xlsx"

    security = _security()
    security.classifications.append(
        SecurityClassification(
            system=GICS,
            nodes={
                ClassificationLevel.LEVEL1: ClassificationNode(value="Information Technology", code="45",),
                ClassificationLevel.LEVEL2: ClassificationNode(value="Software & Services", code="4510",),
                ClassificationLevel.LEVEL3: ClassificationNode(value="Software", code="451030",),
                ClassificationLevel.LEVEL4: ClassificationNode(value="Systems Software", code="45103010",),
            },
        )
    )

    write_excel([security], filename, classifications=ClassificationOutput.GICS,)

    worksheet = _load_worksheet(filename)

    assert _headers(worksheet) == [
        "Name", "Ticker",
        "GICS Sector Code", "GICS Sector",
        "GICS Industry Group Code", "GICS Industry Group",
        "GICS Industry Code", "GICS Industry",
        "GICS Sub-Industry Code", "GICS Sub-Industry",
    ]

    assert worksheet.cell(2, 3).value == "45"
    assert worksheet.cell(2, 4).value == "Information Technology"
    assert worksheet.cell(2, 6).value == "Software & Services"
    assert worksheet.cell(2, 8).value == "Software"
    assert worksheet.cell(2, 10).value == "Systems Software"

    worksheet.parent.close()


def test_write_excel_gecs_and_gics(
    tmp_path: Path,
) -> None:
    """Test simultaneous GECS and GICS output."""

    filename = tmp_path / "output.xlsx"

    security = _security()

    security.classifications.extend(
        [
            SecurityClassification(
                system=GECS,
                nodes={
                    ClassificationLevel.LEVEL1: ClassificationNode(value="Sensitive",),
                    ClassificationLevel.LEVEL2: ClassificationNode(value="Technology",),
                    ClassificationLevel.LEVEL3: ClassificationNode(value="Software",),
                },
            ),
            SecurityClassification(
                system=GICS,
                nodes={
                    ClassificationLevel.LEVEL1: ClassificationNode(value="Information Technology", code="45",),
                    ClassificationLevel.LEVEL2: ClassificationNode(value="Software & Services", code="4510",),
                    ClassificationLevel.LEVEL3: ClassificationNode(value="Software", code="451030",),
                    ClassificationLevel.LEVEL4: ClassificationNode(value="Systems Software", code="45103010",),
                },
            ),
        ]
    )

    write_excel([security], filename, classifications=ClassificationOutput.ALL,)

    worksheet = _load_worksheet(filename)

    headers = _headers(worksheet)

    assert ("GECS Super Sector" in headers)
    assert ("GECS Industry" in headers)
    assert ("GICS Sub-Industry Code" in headers)
    assert ("GICS Sub-Industry" in headers)

    worksheet.parent.close()


def test_write_excel_provider_details_with_classification(
    tmp_path: Path,
) -> None:
    """Test provider details can be output together with classifications."""

    filename = tmp_path / "output.xlsx"

    security = _security()
    security.classifications.append(
        SecurityClassification(
            system=GECS,
            nodes={
                ClassificationLevel.LEVEL1: ClassificationNode(value="Sensitive",),
                ClassificationLevel.LEVEL2: ClassificationNode(value="Technology",),
                ClassificationLevel.LEVEL3: ClassificationNode(value="Software",),
            },
        )
    )
    security.provider_attributes["morningstar"] = {"sector": "Technology"}

    write_excel([security], filename, classifications=ClassificationOutput.GECS, provider_details=True)

    worksheet = _load_worksheet(filename)
    headers = _headers(worksheet)

    assert ("GECS Sector" in headers)
    assert "morningstar.sector" in headers

    classification_column = headers.index("GECS Sector") + 1
    provider_column = headers.index("morningstar.sector") + 1

    assert worksheet.cell(2, classification_column).value == "Technology"
    assert worksheet.cell(2, provider_column).value == "Technology"

    worksheet.parent.close()


def test_write_excel_no_output_selection(
    tmp_path: Path,
) -> None:
    """Test that an empty selection still produces basic security data."""

    filename = tmp_path / "output.xlsx"

    security = _security()

    write_excel([security], filename, classifications=ClassificationOutput.NONE, provider_details=False,)

    worksheet = _load_worksheet(filename)

    assert _headers(worksheet) == ["Name", "Ticker",]

    worksheet.parent.close()
