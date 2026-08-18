"""Tests for processflow input handling."""


# ruff and mypy per file settings
#
# boolean-type arguments
# ruff: noqa: FBT001, FBT002
# others
# ruff: noqa: PLC2701, RUF105

# fmt: off


from pathlib import Path

import openpyxl
import pytest

from equities_classifier.models import SecurityIdentifierType
from equities_classifier.processflow.input import (
    _detect_identifier_type,
    _is_valid_isin,
    read_identifiers,
)


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("US0378331005", True),
        ("DE0007164600", True),
        ("US0378331006", False),
        ("US037833100", False),
        ("US03783310050", False),
        ("0378331005", False),
        ("US0378331005X", False),
        ("", False),
    ],
)
def test_is_valid_isin(
    value: str,
    expected: bool,
) -> None:
    """Test ISIN validation."""

    assert _is_valid_isin(value) is expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("US0378331005", SecurityIdentifierType.ISIN),
        ("DE0007164600", SecurityIdentifierType.ISIN),
        ("AAPL", SecurityIdentifierType.TICKER),
        ("SAP", SecurityIdentifierType.TICKER),
        ("BRK", SecurityIdentifierType.TICKER),
        ("865985", None),
        ("2046251", None),
        ("US0378331006", None),
    ],
)
def test_detect_identifier_type(
    value: str,
    expected: SecurityIdentifierType | None,
) -> None:
    """Test automatic identifier type detection."""

    assert _detect_identifier_type(value) is expected


def _create_input_file(
    path: Path,
    rows: list[tuple[str | None, str]],
) -> None:
    """Create a temporary input workbook."""

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.append(["Type", "Value"])

    for identifier_type, value in rows:
        worksheet.append([identifier_type, value])

    workbook.save(path)
    workbook.close()


def test_read_identifiers_explicit_types(
    tmp_path: Path,
) -> None:
    """Test reading explicitly typed identifiers."""

    filename = tmp_path / "input.xlsx"

    _create_input_file(
        filename,
        [
            ("ISIN", "US0378331005"),
            ("TICKER", "AAPL"),
            ("WKN", "865985"),
        ],
    )

    result = read_identifiers(filename)

    assert [(item.type, item.value) for item in result] == [
        (SecurityIdentifierType.ISIN, "US0378331005"),
        (SecurityIdentifierType.TICKER, "AAPL"),
        (SecurityIdentifierType.WKN, "865985"),
    ]


def test_read_identifiers_automatic_detection(
    tmp_path: Path,
) -> None:
    """Test automatic identifier type detection."""

    filename = tmp_path / "input.xlsx"

    _create_input_file(
        filename,
        [
            (None, "US0378331005"),
            (None, "AAPL"),
        ],
    )

    result = read_identifiers(filename)

    assert [(item.type, item.value) for item in result] == [
        (SecurityIdentifierType.ISIN, "US0378331005"),
        (SecurityIdentifierType.TICKER, "AAPL"),
    ]


def test_read_identifiers_missing_value(
    tmp_path: Path,
) -> None:
    """Test missing identifier value."""

    filename = tmp_path / "input.xlsx"

    _create_input_file(
        filename,
        [
            ("ISIN", ""),
        ],
    )

    with pytest.raises(ValueError, match="Missing identifier value",):
        read_identifiers(filename)


def test_read_identifiers_unknown_type(
    tmp_path: Path,
) -> None:
    """Test unknown explicitly supplied identifier type."""

    filename = tmp_path / "input.xlsx"

    _create_input_file(
        filename,
        [
            ("FOO", "123456"),
        ],
    )

    with pytest.raises(ValueError, match="Unknown security identifier type",):
        read_identifiers(filename)


def test_read_identifiers_unknown_identifier(
    tmp_path: Path,
) -> None:
    """Test identifier which cannot be detected."""

    filename = tmp_path / "input.xlsx"

    _create_input_file(
        filename,
        [
            (None, "865985"),
        ],
    )

    with pytest.raises(ValueError, match="Unable to determine identifier type",):
        read_identifiers(filename)


def test_read_identifiers_without_type_column(
    tmp_path: Path,
) -> None:
    """Test input containing only a Value column."""

    filename = tmp_path / "input.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Value"])
    worksheet.append(["US0378331005"])
    worksheet.append(["AAPL"])
    workbook.save(filename)
    workbook.close()

    result = read_identifiers(filename)

    assert [(item.type, item.value) for item in result] == [
        (SecurityIdentifierType.ISIN, "US0378331005"),
        (SecurityIdentifierType.TICKER, "AAPL"),
    ]


def test_read_identifiers_missing_value_column(
    tmp_path: Path,
) -> None:
    """Test missing Value column."""

    filename = tmp_path / "input.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Type"])
    worksheet.append(["ISIN"])
    workbook.save(filename)
    workbook.close()

    with pytest.raises(ValueError, match="must contain a 'Value' column",):
        read_identifiers(filename)
